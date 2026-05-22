#!/usr/bin/env python3
"""Import the firm's Master Project List (XLSX) + Project Locator (KMZ)
into TaskTrack.

Two source files:

- **`Master List - Numeric ######.xlsx`**: 6,144 projects with one row each.
  Columns (after the header rows on rows 1-4): `Project, Name, Status,
  Start Date, Dormant Date, Client Name, Principal, Component`.

- **`Project Locator.kmz`**: 5,300+ pushpin placemarks in 15 folders
  named by project-number range. Each placemark name is a project number,
  each styleUrl resolves to one of five pushpin colors:
    yellow = primary form placement (legend)
    red    = archived PDF on file
    green  = topo survey on file
    blue   = archived PDF + survey
    pink   = stray legacy
  Four placemarks at the top of the file are the legend itself (names like
  "YELLOW=Project Input Form Placement") and must be skipped.

The script:

  1. Loads both sources.
  2. Normalizes project numbers to the firm-standard `####.##` shape (the
     Excel writes "209.1" for what's really "209.10", and writes plain
     integers like "1014" for what's really "1014.00"). Oddball suffixed
     numbers like "4683.00-SLS - 19" are kept verbatim.
  3. Upserts each Excel row into `projects` keyed on the normalized
     number — every project from the master list lands as a row.
  4. For each KMZ placemark, looks up the parent project (creating a
     blank parent row if the KMZ has a number the master list doesn't),
     replaces that project's `project_sites` rows with the new pin set,
     and updates the legacy `projects.lat`/`projects.lng` columns to the
     primary site's coords so the existing geojson endpoint keeps
     working.
  5. Picks a primary site per project: the yellow pin if any, else the
     first pin in document order.

Idempotent: a second run against the same sources is a no-op aside from
`updated_at` bumps; a re-run against a newer master list cleanly diffs.

Run:
    cd /home/rtoony/projects/collab-tracker
    /home/rtoony/miniconda3/bin/python3 scripts/import_projects_from_master.py \\
        --xlsx "/media/rtoony/13FB-6205/Master List - Numeric 042826.xlsx" \\
        --kmz  "/media/rtoony/13FB-6205/Project Locator.kmz" \\
        --dry-run

    /home/rtoony/miniconda3/bin/python3 scripts/import_projects_from_master.py \\
        --xlsx "/media/rtoony/13FB-6205/Master List - Numeric 042826.xlsx" \\
        --kmz  "/media/rtoony/13FB-6205/Project Locator.kmz"
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))

from sqlalchemy import create_engine, delete, select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.db import DB_PATH  # noqa: E402
from app.models import Project, ProjectSite  # noqa: E402

KML_NS = {"k": "http://www.opengis.net/kml/2.2"}

# Pushpin icon URL -> friendly color name.
_PIN_COLOR_FROM_ICON = {
    "ylw-pushpin.png":  "yellow",
    "red-pushpin.png":  "red",
    "grn-pushpin.png":  "green",
    "blue-pushpin.png": "blue",
    "pink-pushpin.png": "pink",
}

# Pin-color priority for picking the "primary" site of a project. Yellow
# is the canonical "this is the project location" pin per the firm's
# legend; if absent we fall back to the document-order ordering.
_PRIMARY_PRIORITY = {"yellow": 0, "green": 1, "blue": 2, "red": 3, "pink": 4}

# Placemark names that are the KMZ's color legend (the first 4
# placemarks in the original file). They never represent real projects.
_LEGEND_NAME_RE = re.compile(r"^(YELLOW|RED|GREEN|BLUE|PINK)\s*=", re.IGNORECASE)

# Excel master-list header row sits on row 4; data starts at row 5.
_MASTER_LIST_HEADER_ROW = 4
_MASTER_LIST_DATA_START = 5

# Excel Status -> projects.display_status. The master list uses two
# values today (Active / Dormant); be defensive about unexpected ones.
_STATUS_MAP = {
    "active":   "active",
    "dormant":  "dormant",
}


def normalize_project_number(raw) -> str:
    """Canonicalize a project number into the firm's ####.## convention.

    Excel cell values come in as `float | int | str | None`; the KMZ
    delivers strings. Bare integers get `.00` appended. Floats get
    formatted to two decimal places (so 209.1 -> "209.10"). Strings
    that already match `####.##` pass through; anything else (e.g.
    "4683.00-SLS - 19", or completely empty) is returned trimmed
    verbatim so the caller can deal with the oddball.
    """
    if raw is None:
        return ""
    if isinstance(raw, int):
        return f"{raw}.00"
    if isinstance(raw, float):
        return f"{raw:.2f}"
    s = str(raw).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+", s):
        return f"{s}.00"
    if re.fullmatch(r"\d+\.\d", s):
        return f"{s}0"
    if re.fullmatch(r"\d+\.\d{2}", s):
        return s
    if re.fullmatch(r"\d+\.\d{3,}", s):
        # More precision than the firm uses; truncate to two decimals.
        before, after = s.split(".")
        return f"{before}.{after[:2]}"
    return s


def _excel_date_to_iso(val) -> str:
    """Render a possibly-datetime Excel cell into an ISO YYYY-MM-DD string,
    or '' if blank/unparseable."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if not s:
        return ""
    # Best-effort: try a few common shapes; otherwise hand back raw.
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def _read_master_list(path: Path) -> list[dict]:
    """Return a list of dicts, one per Excel data row, with normalized
    project_number and friendly field names."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for the master-list import. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # the master list is single-sheet ("Sheet1")
    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=_MASTER_LIST_DATA_START, values_only=True):
        if not raw or all(v is None for v in raw[:8]):
            continue
        pn, name, status, start, dormant, client, principal, component = raw[:8]
        normalized = normalize_project_number(pn)
        if not normalized:
            # Row with no project number at all — skip silently.
            continue
        rows.append({
            "project_number": normalized,
            "name":           (str(name).strip() if name else ""),
            "status":         _STATUS_MAP.get(
                str(status).strip().lower() if status else "",
                "dormant",
            ),
            "start_date":     _excel_date_to_iso(start),
            "dormant_date":   _excel_date_to_iso(dormant),
            "client":         (str(client).strip() if client else ""),
            "principal":      (str(principal).strip() if principal else ""),
            "component":      (str(component).strip() if component else ""),
        })
    return rows


def _resolve_kmz_styles(root: ET.Element) -> dict[str, str]:
    """Return {style_id: friendly_color_name} resolved through any
    StyleMap indirections. KMZs use StyleMap entries with `normal` and
    `highlight` pair states; we want the `normal` pin appearance."""
    direct: dict[str, str] = {}
    for s in root.findall(".//k:Style", KML_NS):
        sid = s.get("id") or ""
        icon = s.find(".//k:IconStyle/k:Icon/k:href", KML_NS)
        if icon is None or not icon.text:
            continue
        url = icon.text.strip()
        for needle, color in _PIN_COLOR_FROM_ICON.items():
            if needle in url:
                direct[sid] = color
                break

    via_map: dict[str, str] = {}
    for sm in root.findall(".//k:StyleMap", KML_NS):
        sid = sm.get("id") or ""
        for pair in sm.findall("k:Pair", KML_NS):
            key = pair.find("k:key", KML_NS)
            url = pair.find("k:styleUrl", KML_NS)
            if (key is not None and key.text == "normal"
                    and url is not None and url.text):
                ref = url.text.lstrip("#")
                if ref in direct:
                    via_map[sid] = direct[ref]
                break

    return {**direct, **via_map}


def _read_kmz_pins(path: Path) -> list[dict]:
    """Return a list of pin dicts, one per real placemark (legend ones
    filtered out). Each dict carries normalized_pn, lat, lng, pin_color,
    raw_name, and folder_label (the number-range folder)."""
    with zipfile.ZipFile(path) as zf:
        # Most KMZs have a single doc.kml at the root.
        kml_names = [n for n in zf.namelist() if n.lower().endswith(".kml")]
        if not kml_names:
            raise SystemExit(f"no .kml found inside {path}")
        with zf.open(kml_names[0]) as fp:
            tree = ET.parse(fp)
    root = tree.getroot()
    style_color = _resolve_kmz_styles(root)

    out: list[dict] = []
    for folder in root.findall(".//k:Folder", KML_NS):
        fname_el = folder.find("k:name", KML_NS)
        folder_label = (fname_el.text or "").strip() if fname_el is not None else ""
        # The legend lives at the root "Project Locator" folder, alongside
        # the real number-range folders. Real folders have names like
        # "1 - 499", "500 - 999". Skip the root folder's placemarks AND
        # any placemark whose name matches the YELLOW=/RED=/etc. pattern.
        is_legend_folder = folder_label.lower() == "project locator"

        for pm in folder.findall("k:Placemark", KML_NS):
            n_el = pm.find("k:name", KML_NS)
            name_raw = (n_el.text or "").strip() if n_el is not None else ""
            if not name_raw:
                continue
            if _LEGEND_NAME_RE.match(name_raw):
                continue
            if is_legend_folder:
                # Even if a non-legend-named placemark slipped into the
                # "Project Locator" root folder, skip it — folder is
                # reserved for the key.
                continue

            c_el = pm.find(".//k:coordinates", KML_NS)
            if c_el is None or not c_el.text:
                continue
            coord_text = c_el.text.strip().split()[0]  # first vertex only
            try:
                lon_s, lat_s, *_ = coord_text.split(",")
                lng = float(lon_s)
                lat = float(lat_s)
            except (ValueError, IndexError):
                continue

            su = pm.find("k:styleUrl", KML_NS)
            sid = (su.text or "").lstrip("#") if su is not None and su.text else ""
            color = style_color.get(sid, "")

            out.append({
                "raw_name":         name_raw,
                "normalized_pn":    normalize_project_number(name_raw),
                "lat":              lat,
                "lng":              lng,
                "pin_color":        color,
                "folder_label":     folder_label,
            })
    return out


def _pick_primary(pins: list[dict]) -> int:
    """Return the index of the pin that should become `is_primary=1`.
    Yellow wins; failing that, the priority order in _PRIMARY_PRIORITY;
    failing that, the first pin in document order."""
    if not pins:
        return -1
    best_idx, best_rank = 0, _PRIMARY_PRIORITY.get(pins[0].get("pin_color", ""), 99)
    for i, p in enumerate(pins[1:], start=1):
        rank = _PRIMARY_PRIORITY.get(p.get("pin_color", ""), 99)
        if rank < best_rank:
            best_idx, best_rank = i, rank
    return best_idx


def _upsert_project(sess: Session, row: dict, *, now_iso: str = "",
                    from_master: bool = True) -> tuple[Project, str]:
    """Find-or-create a Project on project_number, applying the master-
    list metadata.

    Returns `(project, transition)` where transition is one of:
      "created" — new row inserted from this run
      "updated" — existing row had at least one master-driven field change
      "unchanged" — existing row matched the master exactly; only the
                    `last_seen_in_master_at` stamp moved

    `from_master=True` means this row came from the Excel and so should
    stamp `last_seen_in_master_at` and clear any `vanished_from_master_at`
    flag. `from_master=False` is used by the KMZ-orphan creation path —
    those rows never have their vanish state touched.
    """
    existing = sess.scalar(
        select(Project).where(Project.project_number == row["project_number"])
    )
    if existing is None:
        proj = Project(
            project_number=row["project_number"],
            name=row.get("name", ""),
            client=row.get("client", ""),
            principal=row.get("principal", ""),
            component=row.get("component", ""),
            start_date=row.get("start_date", ""),
            dormant_date=row.get("dormant_date", ""),
            display_status=row.get("status", "dormant"),
            notes=row.get("notes", ""),
        )
        if from_master and now_iso:
            proj.last_seen_in_master_at = now_iso
        sess.add(proj)
        sess.flush()
        return proj, "created"

    # Master always wins. Detect whether any field actually moved so the
    # caller can distinguish a real update from a touch-the-stamp run.
    fields = ("name", "client", "principal", "component",
              "start_date", "dormant_date")
    changed = any(
        row.get(c, getattr(existing, c)) != getattr(existing, c)
        for c in fields
    )
    new_status = row.get("status", existing.display_status)
    if new_status != existing.display_status:
        changed = True

    for c in fields:
        setattr(existing, c, row.get(c, getattr(existing, c)))
    existing.display_status = new_status

    if from_master and now_iso:
        existing.last_seen_in_master_at = now_iso
        # Project re-appeared after having vanished — clear the flag so
        # the row drops out of the vanish report and the admin UI stops
        # highlighting it.
        if existing.vanished_from_master_at:
            existing.vanished_from_master_at = ""
            changed = True

    if changed:
        existing.updated_at = datetime.utcnow()
    return existing, ("updated" if changed else "unchanged")


def _replace_sites_for(sess: Session, proj: Project, pins: list[dict]) -> int:
    """Wipe and re-insert this project's site rows. Returns the number
    of pins inserted. Also mirrors the primary site's coords onto the
    project row so the legacy lat/lng path keeps working."""
    sess.execute(delete(ProjectSite).where(ProjectSite.project_id == proj.id))
    if not pins:
        # Project has no map pins; clear legacy lat/lng so a stale value
        # from an earlier import doesn't ghost on.
        proj.lat = None
        proj.lng = None
        return 0

    primary_idx = _pick_primary(pins)
    for i, p in enumerate(pins):
        sess.add(ProjectSite(
            project_id=proj.id,
            lat=p["lat"],
            lng=p["lng"],
            pin_color=p.get("pin_color", ""),
            raw_name=p.get("raw_name", ""),
            is_primary=1 if i == primary_idx else 0,
            source="kmz",
        ))
    primary = pins[primary_idx]
    proj.lat = primary["lat"]
    proj.lng = primary["lng"]
    return len(pins)


def run_import(xlsx_path: Path, kmz_path: Path, *, db_url: str,
               dry_run: bool, now_iso: str = "") -> dict:
    """Run the importer end-to-end.

    Returns a structured report dict carrying both summary counts and
    per-project transition lists so `sync_master_if_changed.py` can build
    a Telegram digest without re-querying the DB. The report shape is
    versioned via `schema_version` so future changes stay backward-
    compatible.
    """
    excel_rows = _read_master_list(xlsx_path)
    kmz_pins = _read_kmz_pins(kmz_path)

    # Group pins by their normalized project number.
    pins_by_pn: dict[str, list[dict]] = {}
    for p in kmz_pins:
        pins_by_pn.setdefault(p["normalized_pn"], []).append(p)

    excel_pns = {r["project_number"] for r in excel_rows}
    kmz_pns = set(pins_by_pn.keys())

    if not now_iso:
        now_iso = datetime.utcnow().isoformat()

    report = {
        "schema_version":   1,
        "run_at":           now_iso,
        "xlsx_path":        str(xlsx_path),
        "kmz_path":         str(kmz_path),
        "dry_run":          dry_run,
        "excel_rows":               len(excel_rows),
        "kmz_pins_total":           len(kmz_pins),
        "kmz_unique_numbers":       len(kmz_pns),
        "in_both":                  len(excel_pns & kmz_pns),
        "excel_only":               len(excel_pns - kmz_pns),
        "kmz_only_orphans":         len(kmz_pns - excel_pns),
        "projects_upserted":        0,
        "orphan_projects_created":  0,
        "sites_inserted":           0,
        # Transition lists (each is a list of project_numbers — keep them
        # short and stable so the digest can sample/truncate easily).
        "created":          [],
        "updated":          [],
        "unchanged":        [],
        "vanished":         [],   # this-run-only new vanishes
        "returned":         [],   # projects that came back from vanished
    }

    if dry_run:
        return report

    engine = create_engine(db_url, future=True)
    with Session(engine) as sess, sess.begin():
        # 1) Master-list rows -> projects.
        for row in excel_rows:
            existing = sess.scalar(
                select(Project).where(Project.project_number == row["project_number"])
            )
            was_vanished = bool(existing and existing.vanished_from_master_at)
            _, transition = _upsert_project(sess, row, now_iso=now_iso, from_master=True)
            report["projects_upserted"] += 1
            if transition == "created":
                report["created"].append(row["project_number"])
            elif transition == "updated":
                report["updated"].append(row["project_number"])
            else:
                report["unchanged"].append(row["project_number"])
            if was_vanished:
                # _upsert_project cleared the flag; record the return.
                report["returned"].append(row["project_number"])

        # 2) KMZ-only orphans -> projects (blank metadata, raw name in
        # `notes` for traceability).
        for pn in (kmz_pns - excel_pns):
            sample_pin = pins_by_pn[pn][0]
            orphan_row = {
                "project_number": pn,
                "name":           "",
                "status":         "dormant",
                "client":         "",
                "principal":      "",
                "component":      "",
                "start_date":     "",
                "dormant_date":   "",
                "notes":          (
                    "Auto-created from Project Locator KMZ "
                    f"(raw_name={sample_pin['raw_name']!r}, folder="
                    f"{sample_pin['folder_label']!r}); not in master list."
                ),
            }
            _, transition = _upsert_project(sess, orphan_row,
                                            now_iso=now_iso, from_master=False)
            if transition == "created":
                report["orphan_projects_created"] += 1

        sess.flush()

        # 3) Sites: rebuild per project. Fetch each project once by pn.
        projects_by_pn = {
            p.project_number: p
            for p in sess.scalars(select(Project)).all()
        }
        for pn, pins in pins_by_pn.items():
            proj = projects_by_pn.get(pn)
            if proj is None:
                # Shouldn't happen — we created orphans above — but be
                # defensive in case the project number has weird
                # whitespace.
                continue
            report["sites_inserted"] += _replace_sites_for(sess, proj, pins)

        # 4) Vanish detection. A project counts as "newly vanished" if:
        #   - It has been seen in some past master-list run
        #     (`last_seen_in_master_at` is non-empty)
        #   - It was NOT seen in this run (stamp older than now_iso)
        #   - It doesn't already have a `vanished_from_master_at` flag
        #   - Its notes don't mark it as a KMZ-only orphan — those were
        #     never in the master to begin with, so they can't vanish
        #     from it.
        # When found: stamp `vanished_from_master_at = now_iso`, flip
        # `display_status` to 'dormant' (operator-confirmed policy), and
        # record the project_number in the report.
        kmz_orphan_marker = "Auto-created from Project Locator KMZ"
        vanish_candidates = sess.scalars(
            select(Project).where(
                Project.last_seen_in_master_at != "",
                Project.last_seen_in_master_at != now_iso,
                Project.vanished_from_master_at == "",
                ~Project.notes.like(f"%{kmz_orphan_marker}%"),
            )
        ).all()
        for proj in vanish_candidates:
            proj.vanished_from_master_at = now_iso
            proj.display_status = "dormant"
            proj.updated_at = datetime.utcnow()
            report["vanished"].append(proj.project_number)

    return report


# Backward-compat alias for existing tests that read summary-style keys.
# The report dict is a superset of the old stats dict, so callers reading
# old keys keep working transparently.
run_import_stats = run_import


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, type=Path,
                        help="Path to Master List - Numeric.xlsx")
    parser.add_argument("--kmz", required=True, type=Path,
                        help="Path to Project Locator.kmz")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse + report counts, do not touch the DB.")
    parser.add_argument("--db", default=None,
                        help="Override db URL (default: live TaskTrack DB).")
    parser.add_argument("--report-json", default=None, type=Path,
                        help="Write the full structured report to this path "
                             "(used by the sync wrapper to build the Telegram "
                             "digest without re-querying the DB).")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        raise SystemExit(f"missing xlsx: {args.xlsx}")
    if not args.kmz.is_file():
        raise SystemExit(f"missing kmz: {args.kmz}")

    db_url = args.db or f"sqlite:///{DB_PATH}"
    report = run_import(args.xlsx, args.kmz, db_url=db_url, dry_run=args.dry_run)

    if args.report_json:
        args.report_json.parent.mkdir(parents=True, exist_ok=True)
        args.report_json.write_text(json.dumps(report, indent=2, default=str))

    print("=" * 56)
    print("TaskTrack master-list import" + ("  (DRY RUN)" if args.dry_run else ""))
    print("=" * 56)
    print(f"  Excel rows read:                {report['excel_rows']:>6}")
    print(f"  KMZ placemarks read:            {report['kmz_pins_total']:>6}")
    print(f"  KMZ unique project numbers:     {report['kmz_unique_numbers']:>6}")
    print(f"  Numbers in both Excel + KMZ:    {report['in_both']:>6}")
    print(f"  Numbers only in Excel:          {report['excel_only']:>6}")
    print(f"  Numbers only in KMZ (orphans):  {report['kmz_only_orphans']:>6}")
    if not args.dry_run:
        print()
        print(f"  Projects upserted:              {report['projects_upserted']:>6}")
        print(f"  Orphan projects created:        {report['orphan_projects_created']:>6}")
        print(f"  Sites inserted:                 {report['sites_inserted']:>6}")
        print(f"  Created (this run):             {len(report['created']):>6}")
        print(f"  Updated (this run):             {len(report['updated']):>6}")
        print(f"  Unchanged (this run):           {len(report['unchanged']):>6}")
        print(f"  Newly vanished (-> dormant):    {len(report['vanished']):>6}")
        print(f"  Returned from vanished:         {len(report['returned']):>6}")
    print("=" * 56)
    return 0


if __name__ == "__main__":
    sys.exit(main())
